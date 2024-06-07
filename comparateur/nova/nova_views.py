from io import BytesIO
import os
import re
from PyPDF2 import PdfReader
from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import render
import numpy as np
import openpyxl
from decimal import Decimal
from django.http import HttpResponse
import pandas as pd

from comparateur.models import *



def extract_commands(content):
    pattern = r'/ Récapitulatif de Livraison N°: \d+ / Page \d+/\d+'
    content = re.sub(pattern, '', content)
    pattern = r'/ Récapitulatif de Livraison N°: \d+ / Page \d+/\d+\s+'
    content = re.sub(pattern, '', content)
    content = re.sub(r'Récapitulatif(?:\s+\w+){7}', '', content)
    content.replace('"Certaines données personnelles sensibles de vos clients',"")
    commands=content.split("BL")
    return commands
def decortiquer_bl(bl_content):
    produit1 = produit2 = option = None
    bl_lines = bl_content.split('\n')
    
    if len(bl_lines) > 0:
        parts = bl_lines[0].split("du")
        if len(parts) > 1:
            numero_bl = parts[0].strip()
            second_part = parts[1].split('Réf')
            if len(second_part) > 1:
                date = second_part[0].strip()
                porteur = second_part[1].strip()
            else:
                date = porteur = None
        else:
            numero_bl = date = porteur = None
    else:
        numero_bl = date = porteur = None
    
    produit1 = bl_lines[1] if len(bl_lines) > 1 else None
    
    if len(bl_lines) > 2:
        second_line = bl_lines[2].split()
        if len(second_line) > 5:
            produit2 = bl_lines[2]
        else:
            option = bl_lines[2]
    else:
        second_line = []
    
    if len(bl_lines) > 3:
        option = bl_lines[3]
    
    return numero_bl, produit1, produit2, option, date, porteur




def extraire_produit_info(ligne):
    match = re.match(r'(.+?) ((?:[+-]?\d+\.\d+\s*)+) (\d+\.\d+)$', ligne)
    if match:
        # Récupérer les parties correspondantes
        produit = match.group(1).strip()
        corrections = match.group(2).strip().split()
        correction_str = "/".join(corrections)
        prix = match.group(3).strip() if match.group(3) else None
        if prix:
            prix = float(prix)  # Convertir le prix en nombre
            
        return produit, correction_str, prix
    
    else:
        # Si aucune correspondance n'est trouvée, retourner None
        return None, None, None
def trouver_produit_similaire_NovaCell(reference, champ):
    produit_equals = NovaCell.objects.filter(reference=reference.upper())
    
    if produit_equals.exists():
        # Si des produits avec la référence exacte existent, les utiliser directement
        produit_similaire = produit_equals.first()
    else:
        mots_reference = set(reference.upper().split())  # Convertir la référence en majuscules et la diviser en mots
        max_mots_communs = 0  # Initialisez le nombre maximum de mots communs
        produit_similaire = None  # Initialisez la référence du produit similaire

        for produit in NovaCell.objects.all():
            mots_produit = set(produit.reference.upper().replace("-", " ").split())  # Convertir la référence du produit en majuscules et la diviser en mots
            mots_communs = len(mots_reference.intersection(mots_produit))

            # Mettre à jour la référence du produit similaire si le nombre de mots communs est le plus grand
            if mots_communs > max_mots_communs:
                max_mots_communs = mots_communs
                produit_similaire = produit
    
    if produit_similaire:
        prix_produit_similaire = getattr(produit_similaire, champ, 0.00)
    else:
        prix_produit_similaire = None
        
    return produit_similaire, prix_produit_similaire


from thefuzz import fuzz

def nova(request):
    if request.method == 'POST' and request.FILES['pdf_file']:
        products = NovaCell.objects.all()
        product_list = list(products.values())
        df = pd.DataFrame(product_list)
        print(df)

        products = POL.objects.all()
        product_list = list(products.values())
        df_pol = pd.DataFrame(product_list)
        print(df_pol)

        products = AKSESS.objects.all()
        product_list = list(products.values())
        df_aksess = pd.DataFrame(product_list)
        print(df_aksess)

        
        Traitements=["GRAPHEN","MAJOR","SELIS","SEL.INT"]
        pdf_file = request.FILES['pdf_file']
        pdf_data = pdf_file.read()
      
        pdf_buffer = BytesIO(pdf_data)
        pdf_reader = PdfReader(pdf_buffer)
        content = ""
        for page_num in range(0,len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            text = page.extract_text()
            #pattern = r'/ Récapitulatif(?:.*?\b(\w+\s*){7})'
            pattern = r'/ Récapitulatif(?:.*?\b(\w+|\W+\s*){})'
            text = re.sub(pattern, '', text)
            #text = re.sub(r'/\s*Récapitulatif(?:\s+(?:\S+\s*){0,7})?°?', '', text)
            text = re.sub(r'(\d+)\s*/\s*Récapitulatif(?:\s+(?:\S+\s*){0,7})?°?', r'\1', text)
            
            encoded_text = text.encode('utf-8', 'ignore')  # Ignore characters that cannot be encoded
            decoded_text = encoded_text.decode('utf-8')
            content += decoded_text
        #print(content)
        formatted_commands = []
        content.replace('"Certaines données personnelles sensibles de vos clients',"")
        bls = extract_commands(content)
        for bl in bls:
            finding_traitement=[]
           
            
            lines = bl.split('\n')
            if "Certaines données personnelles sensibles" in bl.split('\n')[0] :
                pass
            else:
               
                numero_bl,produit1,produit2,option,date,porteur=decortiquer_bl(bl)
                produit_, correction_, prix_=extraire_produit_info(' '.join(produit1.split()[1:]))
                for t in Traitements:
                    if produit_ and t in produit_:
                        finding_traitement.append(t)
                

                if not finding_traitement:
                     finding_traitement.append("SANSTR")
                if finding_traitement[0]=="SELIS":
                    finding_traitement[0]="SELISXT"
                if finding_traitement[0]=="SEL.INT":
                    finding_traitement[0]="SELISXT"


                   
                

                #print(f"{produit_} and finding is {finding_traitement}")
                if produit_ and "E." in produit_ :
                    produit_=produit_.replace("E.","E")
                if produit_:
                    #produit_=produit_.replace(",",".")
                    produit_=produit_.replace("VERT G","")
                    produit_=produit_.replace("BRUN C","")
                    #produit_=produit_.replace(".,",".")
                    produit_=produit_.replace("E 1","E 1.")
                    produit_=produit_.replace("PERFECT E 1.5","PERFECT E 1.50")
                    produit_=produit_.replace("PERFECT 174","PERFECT 1,74")
                    produit_=produit_.replace('SPH','Sphérique')
                    produit_=produit_.replace('Sphérique.','Sphérique')
                    produit_=produit_.replace('EDEN ZETA 1','EDEN ZETA 1,')
                    produit_=produit_.replace('PERFECT ECO','PERFECT E')
                    
                  
                    
                    
                    #produit_=produit_.replace("PERFECT E 1.6","PERFECT E 1.6")
                    produit_=produit_.replace("E 1.,6","E 1.6")
                    if finding_traitement[0] in produit_:
                        produit_=produit_.split(finding_traitement[0])[0]
                    if "GRIS"  in produit_:
                        produit_=produit_.split("GRIS")[0]
                    if "1 DAY ACUVUE MOIST F" in produit_:
                        produit_=produit_.replace("1 DAY ACUVUE MOIST F","1 DAY ACUVUE MOIST for ASTIGMATISM")

                    
                    
                    
                    produit_=produit_.split("ø")[0]

               

               
                simulaire1,prix_produit_similaire=trouver_produit_similaire_NovaCell(str(produit_),finding_traitement[0])
                if produit2 and len(produit2.split()) > 1:
                    produit__, correction__, prix__=extraire_produit_info(' '.join(produit2.split()[1:]))
                    simulaire2,prix_produit_similaire=trouver_produit_similaire_NovaCell(str(produit__),finding_traitement[0])
                else:
                    produit__= correction__= prix__=simulaire2=None
                #print(f"option= {option}")
                if option :
                    option=option.split()


                if produit_:
                    print("********************")
                   
                    

                    #MAJOR_CLEAN
                    exact_match_row = df[df['reference'] == produit_]
                    if not exact_match_row.empty:
                        simulaire1 = exact_match_row.iloc[0]['reference']
                        prix_produit_similaire = exact_match_row.iloc[0][finding_traitement[0]]
                    else :
                        #df['similarity'] = df['reference'].apply(lambda x: fuzz.ratio(produit_.upper(), x.upper()))
                        df['similarity'] = df['reference'].apply(lambda x: fuzz.ratio(produit_.upper(), x.upper()))
                        max_similarity = df['similarity'].max()
                        rows_with_max_similarity = df[df['similarity'] == max_similarity]
                        tolerance = 5
                        filtered_rows = rows_with_max_similarity[
                        (rows_with_max_similarity[finding_traitement[0]] >= prix_ - tolerance) & 
                        (rows_with_max_similarity[finding_traitement[0]] <= prix_ + tolerance)]
                        if not filtered_rows.empty:
                            # Si des lignes ont été trouvées, obtenir la première ligne (vous pouvez ajuster selon vos besoins)
                            row_with_closest_param = filtered_rows.iloc[0]
                            
                            # Récupérer la valeur de la colonne 'reference'
                            similaire1 = row_with_closest_param['reference']
                            prix_produit_similaire=row_with_closest_param[finding_traitement[0]]
                            
                            print("Produit similaire:", similaire1)
                        #print(rows_with_max_similarity)
                        #simulaire1 = ""
                        
                    
                    if "AKSESS" in produit_:
                        mots = produit_.split()  # Divise la chaîne en mots
                        if len(mots) >= 3:  # Assurez-vous qu'il y a au moins trois mots dans la chaîne
                            mots[1], mots[2] = mots[2], mots[1]
                            produit_ = ' '.join(mots)
                        df_aksess['similarity'] = df_aksess['reference'].apply(lambda x: fuzz.ratio(produit_.upper(), x.upper()))
                        row_with_max_similarity = df_aksess.loc[df_aksess['similarity'].idxmax()]
                        simulaire1 = row_with_max_similarity['reference']
                        prix_produit_similaire=row_with_max_similarity['PRIX']
                        print(f"PRIX {prix_}")
                    if "POL" in produit_:
                        df_pol['similarity'] = df_pol['reference'].apply(lambda x: fuzz.ratio(produit_.upper(), x.upper()))
                        row_with_max_similarity = df_pol.loc[df_pol['similarity'].idxmax()]
                        simulaire1 = row_with_max_similarity['reference']
                        prix_produit_similaire=row_with_max_similarity[finding_traitement[0]]


                    simulaire2=simulaire1
                   
                    print("********************")
                diferrence_1=Decimal('0.00')
                if prix_produit_similaire is not None and prix_ is not None:
                # Remplacer les virgules par des points pour assurer la compatibilité des nombres flottants
                    prix_produit_similaire = str(prix_produit_similaire).replace(",", ".")
                    prix_ = str(prix_).replace(",", ".")

                    # Convertir les chaînes en Decimal pour les calculs
                    prix_produit_similaire_decimal = Decimal(prix_produit_similaire).quantize(Decimal('0.01'))
                    prix_decimal = Decimal(prix_).quantize(Decimal('0.01'))

                    # Calculer la différence
                    diferrence_1 = prix_produit_similaire_decimal - prix_decimal
                if porteur:
                    porteur=porteur.replace(" Prix Facturé H.T.", "")
                formatted_command = {
                "BL": numero_bl,
                "Date": date,
                "Porteur": porteur,
                "Produit1": produit_,
                "Produit2": produit__,
                "Option": option,
                 "Prix1":prix_,
                 "Correction1":correction_,
                 "Prix2":prix__,
                 "Correction2":correction__,
                 "S1":simulaire1,
                 "S2":simulaire2,
                 "S_prix1":prix_produit_similaire,
                 "S_prix2":prix_produit_similaire,
                 'Diff1': diferrence_1.quantize(Decimal('0.01'))



                 
            }

                formatted_commands.append(formatted_command)
            #print(formatted_commands)
        return render(request, 'excel/nova.html', {'formatted_commands': formatted_commands})
                
                
                
    return render(request,'excel/nova.html')






def novadata2(request):
    chemin_fichier_excel = 'C:/Users/Amnzo/Desktop/Nova/nova.xlsx'
    #noms_feuilles = ["1", "2", "3", "4", "5"]  # Liste des noms des feuilles à traiter
    noms_feuilles = ["GLOBALE","POL","AKSESS"]
    
    try:
        # Chargement du classeur Excel
        classeur = openpyxl.load_workbook(chemin_fichier_excel)
        
        # Parcours des feuilles spécifiées
        for nom_feuille in noms_feuilles:
            if nom_feuille in classeur.sheetnames:
                feuille = classeur[nom_feuille]
                
                # Parcours des lignes de la feuille
                for index, ligne in enumerate(feuille.iter_rows(values_only=True)):
                    # Vérifier si la ligne a au moins 8 colonnes (ou le nombre de colonnes attendu)
                    if len(ligne) < 8:
                        print(f"Skipping row {index} in sheet {nom_feuille}: insufficient columns")
                        continue
                    
                    # Transformation des valeurs et gestion des valeurs manquantes
                    def convert_value(value):
                        if value is None or value == '-':
                            return Decimal(0)
                        if isinstance(value, (int, float)):
                            return Decimal(value)
                        value = value.replace("€", "").replace(",", ".").strip()
                        return Decimal(value)
                    
                    try:
                        nova_cell = NovaCell(
                            Classeur_Name=nom_feuille,
                            reference=ligne[0],
                            SANS_TR=convert_value(ligne[1]),
                            DIAPLUS=convert_value(ligne[2]),
                            MAJOR_CLEAN=convert_value(ligne[3]),
                            SELIS_XT=convert_value(ligne[4]),
                            GRAPHENE=convert_value(ligne[5]),
                            AIRLIS_Shock=convert_value(ligne[6]),
                            #=convert_value(ligne[7])
                        )
                        print(nova_cell)
                        nova_cell.save()
                    except Exception as e:
                        print(f"Error saving row {index} in sheet {nom_feuille}: {e}")
                        continue
        
        # Fermeture du classeur
        classeur.close()
        
        return HttpResponse("NOVACELL DATA BASE CREATED")
    
    except FileNotFoundError:
        return HttpResponse("File not found.", status=404)
    
    except KeyError:
        return HttpResponse("Sheet not found.", status=404)
    
    except Exception as e:
        return HttpResponse(f"An error occurred: {str(e)}", status=500)
    




def novadata(request):
    #chemin_fichier_excel = 'C:/Users/Amnzo/Desktop/Nova/nova.xlsx' #LENTILLE
    data_dir = os.path.join(settings.BASE_DIR, 'comparateur', 'data')
    chemin_fichier_excel = os.path.join(data_dir, 'NOVA_LENTILLE.xlsx')
    #chemin_fichier_excel = ''#C:/Users/Amnzo/Desktop/Nova/NOVA_LENTILLE.xlsx
    #noms_feuilles = ["GLOBALE", "POL", "AKSESS"]
    noms_feuilles = ["LENTILLE"]
    
    try:
        # Chargement du classeur Excel
        classeur = openpyxl.load_workbook(chemin_fichier_excel)
        
        # Parcours des feuilles spécifiées
        for nom_feuille in noms_feuilles:
            if nom_feuille in classeur.sheetnames:
                feuille = classeur[nom_feuille]
                
                # Parcours des lignes de la feuille
                for index, ligne in enumerate(feuille.iter_rows(values_only=True)):
                    
                    
                    def convert_value(value):
                        if value is None or value == '-':
                            return Decimal(0)
                        if isinstance(value, (int, float)):
                            return Decimal(value)
                        value = value.replace("€", "").replace(",", ".").strip()
                        return Decimal(value)
                    
                    try:
                        print(nom_feuille)
                        if nom_feuille == "LENTILLE":
                            
                            nova_cell = NovaCell(
                                reference=ligne[0],
                                SANSTR=convert_value(ligne[1]),
                                # DIAPLUS=convert_value(ligne[2]),
                                # MAJOR=convert_value(ligne[3]),
                                # SELISXT=convert_value(ligne[4]),
                                # GRAPHEN=convert_value(ligne[5]),
                                # AIRLISShock=convert_value(ligne[6]),
                                # BLUE_Shock=convert_value(ligne[7]),
                                # Ajoutez les autres colonnes ici
                            )
                            nova_cell.save()
                        elif nom_feuille == "POL":
                            
                            pol = POL(
                                
                                reference=ligne[0],
                                DIAPLUS=convert_value(ligne[1]),
                                ARINTERNE=convert_value(ligne[2]),
                                MAJOR=convert_value(ligne[3]),
                                AIRLIS=convert_value(ligne[4]),
                                # Ajoutez les autres colonnes ici
                            )
                            pol.save()
                        elif nom_feuille == "AKSESS":
                        
                            
                            aksess = AKSESS(

                                reference=ligne[0],
                                PRIX=convert_value(ligne[1]),
                                # Ajoutez les autres colonnes ici
                            )
                            aksess.save()
                    except Exception as e:
                        print(f"Error saving row {index} in sheet {nom_feuille}: {e}")
                        continue
        
        # Fermeture du classeur
        classeur.close()
        
        return HttpResponse("NOVACELL DATA BASE CREATED 06062024")
    
    except FileNotFoundError:
        return HttpResponse("File not found.", status=404)
    
    except KeyError:
        return HttpResponse("Sheet not found.", status=404)
    
    except Exception as e:
        return HttpResponse(f"An error occurred: {str(e)}", status=500)
    

