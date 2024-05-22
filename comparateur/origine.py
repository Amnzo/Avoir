from decimal import Decimal, InvalidOperation
import json
import PyPDF2
from django.conf import settings
import numpy as np
from django.http import JsonResponse,HttpResponse
import re
from django.shortcuts import render,HttpResponse
from .models import Classeur, ExcelData
from django.db.models import Q
# ExcelData.objects.create(reference=data["reference"], UC=data["UC"], HC=data["HC"], ISC=data["ISC"], SCC=data["SCC"], SRC=data["SRC"], SRB=data["SRB"], SRC_UV=data["SRC_UV"], SRBUV=data["SRBUV"], RCC=data["RCC"], price=data["price"])
def extraire_informations_D(texte):
    # Diviser le texte en lignes
    lignes = texte.split('\n')
    # Initialiser les variables pour stocker les informations
    informations_D = []
    # Indicateur pour détecter la ligne contenant uniquement "D" et "G"
    d_found = False
    # Parcourir les lignes du texte
    for ligne in lignes:
        # Vérifier si la ligne contient uniquement "D"
        if ligne.strip() == "D":
            d_found = True
            continue
        # Si "D" est trouvé, ajouter les lignes suivantes jusqu'à trouver "CT"
        if d_found:
            if ligne.strip() != "CT":
                informations_D.append(ligne)
            else:
                break
    return informations_D



def extraire_informations_G(texte):
    # Diviser le texte en lignes
    lignes = texte.split('\n')
    # Initialiser les variables pour stocker les informations
    informations_G = []
    # Indicateur pour détecter la ligne contenant uniquement "D" et "G"
    d_found = False
    # Parcourir les lignes du texte
    for ligne in lignes:
        # Vérifier si la ligne contient uniquement "D"
        if ligne.strip() == "G":
            d_found = True
            continue
        # Si "D" est trouvé, ajouter les lignes suivantes jusqu'à trouver "CT"
        if d_found:
            if ligne.strip() != "CT":
                informations_G.append(ligne)
            else:
                break

   
    return informations_G



    
    

def decortiquer_commande(texte):
    lines = texte.split('\n')
    #print(texte)
    # Initialiser les variables à None

    commande = reference = produit1 = option=None
    commande = lines[1]  # La deuxième ligne est la commande
    option=option1=option2=option3=option4=option5=some1=some2=None
    for i, line in enumerate(lines):
        if line.startswith("Référence"):
            reference = lines[i + 1].strip()
        elif line.startswith("Produit"):
            produit1 = lines[i + 1].strip()
        elif line.startswith("Options"):
            option=f"{lines[i + 1].strip()}"
            option1 = f"{lines[i + 2].strip()}"
            option2 = f"{lines[i + 3].strip()}"
            option3=f"{lines[i + 4].strip()}"
           
            if i + 5 < len(lines):
                option4 = lines[i + 5].strip()
            if i + 6 < len(lines):
                option5 = lines[i + 6].strip()
           # Decimal(str(prix_d).replace(',', '.'))
            some1 = Decimal(str(option1).replace(',', '.'))-Decimal(str(option2).replace(',', '.'))
            if option4 is not None and option5 is not None:
                try:
                    option4_decimal = Decimal(option4.replace(',', '.'))
                    option5_decimal = Decimal(option5.replace(',', '.'))
                    some2 = option4_decimal - option5_decimal
                except InvalidOperation:
                    some2 = Decimal(0)
            else:
                some2 = Decimal(0)
            
            
           
    return commande, reference, produit1,option,option1,option2,option3,option4,option5,some1,some2



import openpyxl
def data(request):
    chemin_fichier_excel = 'C:/Users/Amnzo/Desktop/jonathan/rania.xlsx'
    nom_feuille = 'SEIKO Offre 2e paire'
    classeur_parent=Classeur.objects.get(pk=3)

    classeur = openpyxl.load_workbook(chemin_fichier_excel)
    feuille = classeur[nom_feuille]
    # Parcourir les lignes de la feuille de calcul et créer les objets ExcelData
    i=0
    for ligne in feuille.iter_rows(values_only=True):
        i=i+1
        print("ligne:", ligne)  # Add this line for debugging
        p=ExcelData()
        p.classeur=classeur_parent
        p.reference=ligne[0]
        #p.HSC=Decimal(ligne[4]) if ligne[4] is not None else Decimal(0)
        # p.UC=Decimal(ligne[1]) if ligne[1] is not None else Decimal(0)
        # p.HC=Decimal(ligne[2]) if ligne[2] is not None else Decimal(0)
        # p.ISC=Decimal(ligne[3]) if ligne[3] is not None else Decimal(0)
        # p.SCC=Decimal(ligne[4]) if ligne[4] is not None else Decimal(0)
        # p.SRC=Decimal(ligne[5]) if ligne[5] is not None else Decimal(0)
        # p.SRB=Decimal(ligne[6]) if ligne[6] is not None else Decimal(0)
        # p.SRCUV=Decimal(ligne[7]) if ligne[7] is not None else Decimal(0)
        # p.SRBUV=Decimal(ligne[8]) if ligne[8] is not None else Decimal(0)
        # p.RCC=Decimal(ligne[9]) if ligne[9] is not None else Decimal(0)
        #seiko 2PAIRE
        p.UC=Decimal(ligne[1]) if ligne[1] is not None else Decimal(0)
        p.HC=Decimal(ligne[2]) if ligne[2] is not None else Decimal(0)
        p.SCC=Decimal(ligne[3]) if ligne[3] is not None else Decimal(0)
        p.SUNUC=Decimal(ligne[4]) if ligne[4] is not None else Decimal(0)
        p.SUNHC=Decimal(ligne[5]) if ligne[5] is not None else Decimal(0)
        p.SUNISC=Decimal(ligne[6]) if ligne[6] is not None else Decimal(0)
        p.POLA_UC=Decimal(ligne[7]) if ligne[7] is not None else Decimal(0)
        p.POLA_ISC=Decimal(ligne[8]) if ligne[8] is not None else Decimal(0)


        #DERNIER CLASSEUR Game
        p.prix=Decimal(ligne[1]) if ligne[1] is not None else Decimal(0)
        p.save()
        print(f"{p}")
    classeur.close()
    
    return HttpResponse("DATA BASE CREATED")

from django.db.models import CharField, Value as V
from django.db.models.functions import Length



    #return produits_similaires
from Levenshtein import distance
def trouver_produit_similaire2(reference):
    produits = ExcelData.objects.all()
    distances = [(produit.reference, distance(reference, produit.reference)) for produit in produits]
    distances.sort(key=lambda x: x[1])
    print(distances[0][0])
    return distances[0][0]
def trouver_produit_similaire(reference):
    if "JETSTAR" in reference:
        reference=trouver_produit_similaire2(reference)

    mots_reference = reference.upper().split()  # Convertir la référence en majuscules et la diviser en mots
    #print(mots_reference)
    max_mots_communs = 0  # Initialisez le nombre maximum de mots communs
    produit_similaire = None  # Initialisez la référence du produit similaire
    # Parcourir tous les produits dans la base de données
    for produit in ExcelData.objects.all():
        mots_produit = produit.reference.upper()  # Convertir la référence du produit en majuscules et la diviser en mots
        #print(mots_produit)
        # Compter le nombre de mots communs entre la référence donnée et la référence du produit
        mots_communs = 0
        for mot in mots_reference:
            if mot in mots_produit:
                mots_communs += 1
        
        # Vérifier si le produit contient "kids" et la référence donnée ne contient pas "kids"
        if "KIDS" in produit.reference.upper() and "KIDS" not in reference.upper():
            continue  # Passer au produit suivant
        
        # Mettre à jour la référence du produit similaire si le nombre de mots communs est le plus grand
        if mots_communs > max_mots_communs:
            max_mots_communs = mots_communs
           
            produit_similaire = produit.reference

    return produit_similaire






import re
def is_word_in_string(word, string):
    pattern = re.compile(r'\b{}\b'.format(re.escape(word)), re.IGNORECASE)
    return re.search(pattern, string) is not None


def extraire_prix2(texte):
    lines = texte.split("\n")
    dernier_decimal_apres_ET = []
    ligne_precedente_ET = False
    
    for i, ligne in enumerate(lines):
        if ligne.startswith("ET"):
            parties = ligne.split()
            derniers_decimaux = []
            for j in range(i + 1, len(lines)):
                parties_suivantes = lines[j].split()
                for partie in reversed(parties_suivantes):
                    if partie.replace(',', '').replace('.', '').isdigit():
                        dernier_dec = partie.replace(',', '.')  # Assurez-vous que la virgule est au format décimal
                        derniers_decimaux.append(dernier_dec)
                        break
                else:
                    # Si aucun nombre n'est trouvé dans cette ligne, arrêtez la boucle
                    break
            dernier_decimal_apres_ET.append(derniers_decimaux)
    #print(dernier_decimal_apres_ET)
              
    return dernier_decimal_apres_ET

# def extraire_prix(texte):
#     lines = texte.split("\n")
#     #print(lines)
#     dernier_decimal_apres_ET = []
#     ligne_precedente_ET = False
    
#     for i, ligne in enumerate(lines):
#         if ligne.startswith("ET"):
#             ligne_precedente_ET = True
#         elif ligne_precedente_ET:
#             parties = ligne.split()
#             dernier_dec = None
#             for j in range(i, len(lines)):
#                 parties_suivantes = lines[j].split()
#                 for partie in reversed(parties_suivantes):
#                     if partie.replace(',', '').replace('.', '').isdigit():
#                         dernier_dec = partie
#                         break
#                 if dernier_dec:
#                     break
#             if dernier_dec:
#                 dernier_dec = dernier_dec.replace(',', '.')  # Assurez-vous que la virgule est au format décimal
#                 dernier_decimal_apres_ET.append(dernier_dec)
#             else:
#                 dernier_decimal_apres_ET.append("Pas de décimal trouvé")
#             ligne_precedente_ET = False
              
#     return dernier_decimal_apres_ET


def analyser_commande(commande):
    #prices = extraire_prix(commande)
    prices_ = extraire_prix2(commande)
    #print(prices_)
    is_d = is_g = False

    for cmd in commande.split('\n'):
        # Vérifier si la ligne contient uniquement "D" ou "G"
        if cmd.strip() == "D":
            is_d = True
        if cmd.strip() == "G":
            is_g = True

    prix_d = prix_g = 0

    if len(prices_) == 2:
        prix_d = prices_[0][-1]
        prix_g = prices_[1][-1]
        #print(f" both {is_d} - {is_g} {prices}")

    elif len(prices_) == 1:
        if is_d:
            prix_d = prices_[0][-1]
            #print(f" D {is_d} - {is_g} {prices}")
        elif is_g:
            prix_g = prices_[0][-1]
            #print(f" G {is_d} - {is_g} {prices}")

    #print(f"Prix pour D: {prix_d}, Prix pour G: {prix_g}")
    return prix_d,prix_g




def extract_commands(content): 
    commands=content.split("Commande")
    return commands

from django.db.models import DecimalField  
from PyPDF2 import PdfReader
from io import BytesIO
def read_pdf(request):
    # Path to the PDF file
    #pdf_path = 'C:/Users/Amnzo/Desktop/jonathan/6.pdf'
    if request.method == 'POST' and request.FILES['pdf_file']:
        pdf_file = request.FILES['pdf_file']
        pdf_data = pdf_file.read()
        pdf_buffer = BytesIO(pdf_data)
        # Create a PdfReader object using the BytesIO buffer
        pdf_reader = PdfReader(pdf_buffer)
        content = ""
            #print(f"Nombre de page = {pdf_reader.pages}")
        for page_num in range(2,len(pdf_reader.pages)):
                #print(page_num)
            page = pdf_reader.pages[page_num]
            text = page.extract_text()
            encoded_text = text.encode('utf-8', 'ignore')  # Ignore characters that cannot be encoded
            decoded_text = encoded_text.decode('utf-8')
            content += decoded_text
        commands = extract_commands(content)
        formatted_commands = []
        for command in commands[1:]:
          
            commande_decortiquer,reference_decortiquer, produit_1_decortiquer,option,option1,option2,option3,option4,option5,some1,some2=decortiquer_commande(command)
            #print("Option")
            #print(option)
            D_decortiquer=extraire_informations_D(command)
            G_decortiquer=extraire_informations_G(command)
            prix_d,prix_g=analyser_commande(command)
            extraire_prix2(command)
            
           
            champs_excel_data = [field.name for field in ExcelData._meta.get_fields()]            
            champs_decimal = [field for field in champs_excel_data if isinstance(ExcelData._meta.get_field(field), DecimalField)]
            valuer=["UC"]
            valuer_2=["UC"]
            for champ in champs_decimal:
                if is_word_in_string(champ.upper(), produit_1_decortiquer.upper()):
                    valuer[0]=champ
                else:
                    pass

            for champ in champs_decimal:
                if is_word_in_string(champ.upper(), produit_1_decortiquer.upper()):
                    #valuer_2.append(champ)
                    valuer_2[0]=champ
                else:
                    pass

            produit_similaire_1 = trouver_produit_similaire(produit_1_decortiquer.replace('JET STAR','JETSTAR').replace('85%','')) 
            produit_similaire_2 = trouver_produit_similaire(produit_1_decortiquer.replace('JET STAR','JETSTAR').replace('85%','')) 
            product_1= ExcelData.objects.filter(reference=produit_similaire_1).first()
            product_2= ExcelData.objects.filter(reference=produit_similaire_2).first()
            prix_catalogue_1 = None
            prix_catalogue_2 = None
            if product_1:
                for champ in reversed(valuer):
                    if hasattr(product_1, champ):
                        prix_catalogue_1 = getattr(product_1, champ)
                        break
            if  product_1 and product_1.reference.startswith('CLASSE A'):
                
                valuer[0]="HSC"
                print(valuer)
                print(product_1)
                print(product_1.HSC)
                prix_catalogue_1=Decimal(product_1.HSC).quantize(Decimal('0.01'))
            startvision=["STARVISION","JETSTAR","SENSO"]      
            if  product_1 and any(product_1.reference.startswith(start) for start in startvision):
                print("***************************************************")
                valuer[0]="PRIX"
                print(valuer)
                print(product_1)
                print(product_1.prix)
                prix_catalogue_1=Decimal(product_1.prix).quantize(Decimal('0.01'))  
            #if product_1 and "non traité" in produit_1_decortiquer and not  product_1.classeur.nom=='SEIKO Classe A':
                #print("***************************************************")
                #print("non triater trouver----------------------------------")
                #print(produit_1_decortiquer)
                #valuer.append("UC")
                #print(valuer)
                #print(product_2)
                #print(product_2.UC)
                #print("***************************************************")
                #prix_catalogue_1=product_1.UC.quantize(Decimal('0.01'))
                #print(prix_catalogue_1) 
            if product_2 and "non traité" in produit_1_decortiquer and not product_2.classeur.nom=='SEIKO Classe A':
                #print("non triater trouver----------------------------------")
                #valuer.append("UC")
                prix_catalogue_2=product_2.UC.quantize(Decimal('0.01'))
                #print(prix_catalogue_2) 
            


            #-------------------------

            if product_2:
                for champ in reversed(valuer):
                    if hasattr(product_2, champ):
                        prix_catalogue_2 = getattr(product_2, champ)
                        break  
            if  product_2 and product_2.classeur.nom=='SEIKO Classe A':
                valuer.append("HSC")
                prix_catalogue_2=Decimal(product_2.HSC).quantize(Decimal('0.01'))      
            if  product_2 and product_2.classeur.nom=='SEIKO Gamme ECO':
                valuer.append("PRIX ")
                prix_catalogue_2=Decimal(product_2.prix).quantize(Decimal('0.01'))  
                #print(f"prix_catalogue_2 {prix_catalogue_2}")
            

           
          
            if product_1:
                remise_entier_1=remise_produit_1=product_1.remise
            if product_1:
                remise_entier_2=remise_produit_2=product_2.remise
            
            
            parties = str(remise_produit_1).split(".")
            if len(parties) > 1 and int(parties[1]) == 0:
                remise_entier = int(parties[0])
            else:
                remise_entier = remise_produit_1
            taux_remise_decimal_d =Decimal(str(prix_d).replace(',', '.'))-((Decimal(remise_entier) / Decimal(100))*Decimal(str(prix_d).replace(',', '.')))
            taux_remise_decimal_g =Decimal(str(prix_g).replace(',', '.'))-((Decimal(remise_entier) / Decimal(100))*Decimal(str(prix_g).replace(',', '.')))
            #taux_remise_decimal = Decimal(remise_entier) / Decimal(100)  # Calcul du taux de remise en décimal
            #montant_remise = Decimal(str(prix_d).replace(',', '.')) - taux_remise_decimal

            Diff_d=Diff_g = 0        
            try:
                Diff_d = (prix_catalogue_1-taux_remise_decimal_d ).quantize(Decimal('0.01'))
                #result = str(Diff)  # Convert Diff back to string for return
            except TypeError:
                Diff_d = 0
            try:
                Diff_g = (prix_catalogue_2-taux_remise_decimal_g).quantize(Decimal('0.01'))
                #result = str(Diff)  # Convert Diff back to string for return
            except TypeError:
                Diff_g = 0
            
           
            formatted_command = {
                
                "Commande": commande_decortiquer ,#command.split("|")[0] , #.split("|")[0],
                "Référence": reference_decortiquer,
                "Produit_1": produit_1_decortiquer,
                "Produit_2_": produit_1_decortiquer,
                "CorrectionD":D_decortiquer,
                "CorrectionG":G_decortiquer, 
                "Similaire_1": produit_similaire_1,
                "Similaire_1": produit_similaire_2,
                "Remise" : remise_entier,
                "Apres_Remise_d" : taux_remise_decimal_d.quantize(Decimal('0.01')),
                "Apres_Remise_g" : taux_remise_decimal_g.quantize(Decimal('0.01')),
                "Valeur" : valuer[-1],
                "Prix_catalogue_1" :prix_catalogue_1,
                "Prix_catalogue_2" :prix_catalogue_2,
                "Product_1":product_1,
                "Product_2":product_2,
                "Prix_Facture_D": prix_d,   #re.findall(r'\d+,\d+', command[3].strip())[-1]
                "Prix_Facture_G": prix_g,   #re.findall(r'\d+,\d+', command[3].strip())[-1]
                "Diff_d":Diff_d,
                "Diff_g":Diff_g,
                "Option":option,
                "Option1":option1,
                "Option2":option2,
                "Option3":option3,
                "Option4":option4,
                "Option5":option5,
                "Some1":some1,
                "Some2":some2

            
                
                
            }
            formatted_commands.append(formatted_command)

        return render(request, 'excel/excel.html', {'formatted_commands': formatted_commands})
    return render(request, 'excel/excel.html')
    




import random
from decimal import Decimal
from .models import ExcelData

def generate_random_decimal():
    # Générer un nombre décimal aléatoire entre 0 et 1000 avec 2 décimales
    return Decimal(random.uniform(0, 1000)).quantize(Decimal('0.01'))

def generate_random_values(request):
    # Récupérer tous les objets ExcelData
    produits = ExcelData.objects.all()

    # Parcourir chaque produit et attribuer des valeurs décimales aléatoires
    for produit in produits:
        # Attribuer des valeurs décimales aléatoires à chaque champ DecimalField
        produit.UC = generate_random_decimal()
        produit.HC = generate_random_decimal()
        produit.ISC = generate_random_decimal()
        produit.SCC = generate_random_decimal()
        produit.SRC = generate_random_decimal()
        produit.SRB = generate_random_decimal()
        produit.SRCUV = generate_random_decimal()
        produit.SRBUV = generate_random_decimal()
        produit.RCC = generate_random_decimal()
        produit.SUNUC = generate_random_decimal()
        produit.SUNHC = generate_random_decimal()
        produit.SUNSCC = generate_random_decimal()
        produit.SUNUC = generate_random_decimal()
        produit.SUNUC = generate_random_decimal()
        produit.POLA_UC = generate_random_decimal()
        produit.POLA_ISC = generate_random_decimal()
        produit.prix = generate_random_decimal()

        # Sauvegarder les modifications dans la base de données
        produit.save()

    #print("Valeurs décimales aléatoires générées avec succès pour tous les produits.")
    return HttpResponse("*********************")




def delete(request):
    products_to_delete = ExcelData.objects.filter(classeur__nom='SEIKO Gamme ECO')
    # Delete the products
    products_to_delete.delete()
    return HttpResponse("products_to_delete")




from django.http import HttpResponse
from openpyxl import Workbook

from bs4 import BeautifulSoup
import io

from openpyxl import Workbook
from django.http import HttpResponse

# Fonction pour exporter les données vers Excel
from openpyxl import Workbook
from django.http import HttpResponse
from bs4 import BeautifulSoup

# Fonction pour extraire les données du tableau HTML
# Fonction pour extraire les données du tableau HTML
def extract_table_data(html_data):
    # Initialiser une liste pour stocker les données du tableau
    table_data = []
    # Utiliser BeautifulSoup pour parser le HTML
    soup = BeautifulSoup(html_data, 'html.parser')

    # Trouver le premier tableau dans le HTML
    table = soup.find('table')

    # Si un tableau est trouvé
    if table:
        # Trouver toutes les lignes dans le tableau
        rows = table.find_all('tr')

        # Pour chaque ligne dans le tableau
        for row in rows:
            # Trouver toutes les cellules dans la ligne
            cells = row.find_all(['td', 'th'])

            # Extraire le texte de chaque cellule et ajouter à la liste des données du tableau
            row_data = [cell.get_text(strip=True) for cell in cells]

            # Vérifier si la colonne "Correction" est présente et la nettoyer
            if "Correction" in row_data:
                correction_index = row_data.index("Correction") + 1
                row_data[correction_index] = row_data[correction_index].strip()

            table_data.append(row_data)

    return table_data


# Fonction pour exporter les données vers Excel
def export_to_excel(request):
    # Récupérer les données du formulaire POST
    table_data = request.POST.get('table_data')

    # Extraire les données du tableau HTML
    data = extract_table_data(table_data)

    # Créer un nouveau classeur Excel
    wb = Workbook()
    ws = wb.active

    # Ajouter les données au classeur Excel
    for row in data:
        ws.append(row)

    # Définir la largeur de colonne pour chaque colonne
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Enregistrer le classeur Excel dans un objet BytesIO
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Retourner une réponse de téléchargement du fichier Excel
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="commandes.xlsx"'
    return response

import pandas as pd
import os
import pandas as pd
import re
import pandas as pd
import re
def database():
    #excel_file = "C:/Users/Amnzo/Desktop/ophtalmic/BL/DB.xlsx"
    data_dir = os.path.join(settings.BASE_DIR, 'comparateur', 'data')
    excel_file = os.path.join(data_dir, 'BAYA.xlsx')
    dfxl = pd.read_excel(excel_file, header=None)
    lignes_valides = []
    
    for index, ligne in dfxl.iterrows():
        #print(ligne)
        if not ligne.isnull().all():
            if isinstance(ligne[1], str) and re.match(r'^\d+,\d+ € \d+,\d+% \d+,\d+ €$', ligne[1]):
                print(ligne)
                nom = ligne[0]
                valeur = ligne[1]
                nom_sans_premier_mot = ' '.join(nom.split(' ')[1:])
                match = re.match(r'^(\d+,\d+) € (\d+,\d+)% (\d+,\d+) €$', valeur)
                if match:
                    prix_net = match.group(3).replace(',', '.')              
                    lignes_valides.append([nom_sans_premier_mot, prix_net])
    
    df_resultat = pd.DataFrame(lignes_valides, columns=['Nom','Prix_Net'])
    df_resultat['Prix_Net'] = df_resultat['Prix_Net'].astype(float)
    print(df_resultat)
    return df_resultat


def database2():
    excel_file = "C:/Users/Amnzo/Desktop/ophtalmic/BL/DB2.xlsx"
    dfxl = pd.read_excel(excel_file, header=None)
    print(dfxl)
    


def database2():
    excel_file = "C:/Users/Amnzo/Desktop/ophtalmic/BL/DB2.xlsx"
    dfxl = pd.read_excel(excel_file, header=None)
    
    # Filtrer et extraire les lignes valides
    lignes_valides = []
    for index, ligne in dfxl.iterrows():
        if not ligne.isnull().all() and index > 2:  # Ignorer les 3 premières lignes
            nom = ligne[0]
            conditionnement = ligne[1]
            prix = ligne[2]
           
            if isinstance(nom, str) and isinstance(conditionnement, str):
                    if "Flacon" in conditionnement or "Boite de" in conditionnement:
                        #prix_net = float(prix.replace(',', '.').replace(' €', ''))
                        lignes_valides.append([nom, prix])
    
    # Créer le DataFrame final
    df_resultat = pd.DataFrame(lignes_valides, columns=['Nom', 'Prix_Net'])
    #print(df_resultat)
    return df_resultat
           



def extract_commands_ophtal(content):
    # Séparer les commandes à partir de la chaîne "B.L. No"
    commands = content.split("du ")
    return commands



import pdfplumber
import re
from datetime import datetime
import pdfplumber
import re
import json

import pdfplumber
import pdfplumber

header_identifiers = [
    "Extrait des Conditions Générales de Vente "
    "FACTURATION",
    "92700 COLOMBES",
    "38 RUE ST DENIS",
    "OPTIC 2000 SCORTIS",
    "LIVRAISON",
    "Code adhérent :",
    "Page",
    "*0147853314*",
    "FR75479118911",
    "FACTURE FC22000459",
    "Type facture Numero Client Mode de reglement Echeance",
    "Désignation Qté PU HT PU Net HT Total HT",
    "Référence commande client",
    "%Rem",
    "Virement 30 jours fin de mois",
    "Total Lentilles+Pack: 2",
    "ESCOMPTE POUR REGLEMENT COMPTANT : 0%",
    "PENALITES DE RETARD AU TAUX DE 12% AN",
    "INDEMNITES DE RETARD POUR FRAIS DE RECOUVREMENT 40€ (quarante euros)",
    "Depuis le 25/05/18 et le RGPD, nous vous informons de la mise à jour de nos CGV consultables sur oph",
    "Joindre la réference @#08821|0147853314|1|STD|O03|0|FRA|Depuis le 25/05/18 et le RGPD, nous vous informons de la mise à jour de nos CGV consultable",
    "FC22000459 à votre facture",
]

additional_ignore_text = "Extrait des Conditions Générales de Vente"
stop_text = "Base HT Total remise(s) Total net HT Port"
import pdfplumber

def extract_text_from_pdf(pdf_path):
    header_identifiers = ["Header1", "Header2"]  # Remplacez par vos identifiants de header
    additional_ignore_text = "Text to ignore"  # Remplacez par le texte à ignorer
    stop_text = "Text to stop processing"  # Remplacez par le texte qui arrête le traitement de la page
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages[:-1]:
                page_text = page.extract_text()
                #OPTIC 2000 SCORTIS OPTIC 2000 SCORTIS 38 RUE ST DENIS 38 RUE ST DENIS 92700 COLOMBES 92700 COLOMBES

                if "Extrait des Conditions Générale de Vente de la Société Ophtalmic Compagnie" in page_text:
                    continue  # Passe à la page suivante si le texte est trouvé
                if "Base HT Total remise(s) Total net HT Port Total HT Taux TVA Mt TVA Total à payer" in page_text:
                    continue
                page_text=page_text.replace("LIVRAISON FACTURATION OPTIC 2000 SCORTIS OPTIC 2000 SCORTIS 38 RUE ST DENIS 38 RUE ST DENIS 92700 COLOMBES 92700 COLOMBES ",'')
                page_text=page_text.replace("LIVRAISON FACTURATION OPTIC 2000 SCORTIS OPTIC 2000 SCORTIS 38 RUE ST DENIS 38 RUE ST DENIS 92700 COLOMBES 92700 COLOMBES",'')
                page_text = re.sub(r"(Extrait des Conditions Générales de Vente de la Société Ophtalmic Compagnie.*?HT Total HT 1 2 0147853314 OPTIC 2000 SCORTIS)", "", page_text, flags=re.DOTALL)
                
                lines = page_text.split('\n')          
                ignore_rest_of_page = False
                for line in lines:
                    if any(header in line for header in header_identifiers):
                        continue
                    if additional_ignore_text in line:
                        continue
                    if stop_text in line:
                        ignore_rest_of_page = True
                    if ignore_rest_of_page:
                        continue
                    text += line + "\n"
            #print(text)
            return text
    except Exception as e:
        print(f"An error occurred: {e}")
        return ""

# Utilisation de la fonction avec un chemin vers le fichier PDF
pdf_path = 'path_to_your_pdf_file.pdf'
extract_text_from_pdf(pdf_path)

# Utilisation de la fonction avec un chemin vers le fichier PDF
pdf_path = 'path_to_your_pdf_file.pdf'
extract_text_from_pdf(pdf_path)

import os
import re
import pandas as pd
from fuzzywuzzy import fuzz
from django.conf import settings
def supprimer_premier_mot(chaine):
    return ' '.join(chaine.split()[1:]) if pd.notnull(chaine) else chaine

def parse_invoice_text(raw_text):

    invoices = []
    current_invoice = {}
    lines = raw_text.split('\n')
    dfs = []
    data_dir = os.path.join(settings.BASE_DIR, 'comparateur', 'data')
    filenames = ['BL_2021.csv', 'BL_2022.csv', 'BL_2023.csv']

    # Chargement des fichiers CSV
    for filename in filenames:
        file_path = os.path.join(data_dir, filename)
        df = pd.read_csv(file_path, encoding='latin1', delimiter=';')
        dfs.append(df)

    # Concaténation des dataframes
    df_combined = pd.concat(dfs, ignore_index=True)

    # Chargement du fichier Excel
    excel_file = os.path.join(data_dir, 'LAST.xlsx')  #AMNZOOO.xlsx
    #excel_file2 = os.path.join(data_dir, 'AMNZO.xlsx')
    dataframe = pd.read_excel(excel_file,names=['Nom', 'Prix_Net'])
    
    print("DataFrame avant modification:")
    print(dataframe)
    #dataframe['Nom'] = dataframe['Nom'].apply(supprimer_premier_mot)
    #dataframe['Nom_Modifie'] = dataframe['Nom'].apply(supprimer_premier_mot)
    #print(dataframe)
    #dataframe[dataframe['Nom'].str.contains('OPTIMISATION', na=False)]
    #print("DataFrame après modification:")
    #print(new_data_frame)
    #dataframe['Nom'] = dataframe['Nom'].apply(lambda x: ' '.join(x.split()[1:]))
    no_match=[]


    for line in lines:
        # Nettoyage de la ligne
        line = line.strip()
        print(" ")
        print(" ")
        print(" ")
        #print(line)
        
        # Vérification du numéro de BL
        if re.match(r'B.L. No \d{2}-\d{8} du \d{2}/\d{2}/\d{4}', line):
            # Ajout de la facture courante à la liste des factures
            if current_invoice:
                invoices.append(current_invoice)
            current_invoice = {'BLNo': line.strip(), 'Products': []}
        elif '***** Port :' in line:
            # Extraction du port
            current_invoice['Port'] = line.split(':')[-1].strip()
        elif re.search(r'\b\d+,\d+\b|\b\d+\.\d+\b', line):
            # Extraction des informations sur le produit
            parts = line.split()
            print(f"'BLNo': {line.strip()}")
            print(parts)
            if len(parts) >= 5:
                description = ' '.join(parts[:-5]).strip()
                discount = parts[-1].strip()
                # Garder le produit uniquement si la remise est "0,00" ou une valeur décimale
                if discount == "0,00" or re.match(r'^\d+\.\d+$', discount):
                    # Recherche du produit dans le DataFrame combiné
                    resultat = df_combined[df_combined['No livraison'] == current_invoice['BLNo'].split()[2]]
                    if not resultat.empty:
                        # Liste pour stocker les designations_from_bl
                        designations_from_bl = []
                        for index, row in resultat.iterrows():       
                            porteur = row.iloc[8]
                            designation_from_bl = row.iloc[11]
                            designations_from_bl.append(designation_from_bl)  # Stocker les designations dans la liste
                            try:
                                if pd.isna(row.iloc[12]):
                                    qtt = 1
                                else:
                                    qtt = int(row.iloc[12])
                            except ValueError:
                                qtt = 1 

                        diff_ophtal= 0 

                        #description.replace("/","").replace()   
                       

                        match = re.search(r'^(.*?)(?=\b[A-Z]+/\d+/[A-Z]+\b)', description)
                        if match:
                            description=match.group(1)
                            
                        
                        match = re.search(r'^(.*?)(?=[A-Z]+/\d+/[A-Z\s]+)', description)

                        if match:
                            
                            description=match.group(1).strip()
                            
                        pattern = re.compile(r'^(.*?)(?=\b[A-Z]+/\d+/\S+)')
                        match = pattern.search(description)
                        if match:
                            description=match.group(1).strip()

                        
                        new_designation=description
                        new_designation=new_designation.replace("Essai","").replace('VERT','').replace('GRIS','').replace("Diam","")
                        match_des = re.search(r'[\+\-]\d', new_designation)
                        if match_des:
                            new_designation= new_designation[:match_des.start()].strip()
                        
                        print(new_designation)
                        
                        
                        
                        # match_1 = re.match(r'^[A-Z]+\s?\d+\s?[A-Z]+\s?\d+', new_designation)
                        # verified =False
                        # if match_1:
                        #     print("MATCH 1")
                        #     #print(new_designation)
                        #     new_designation=match_1.group()
                        #     verified=True
                        # match_2=re.match(r'^[A-Z]+\s?\d+\s?[A-Z]+\s?\d+', new_designation)
                        # if match_2 and not verified :
                        #     print("MATCH 2")
                        #     #print(new_designation)
                        #     new_designation=match_2.group()
                        #     verified=True
                        # #match_3 = re.search(r'^[A-Z]+\s?\d*[A-Z]+\s?\d*', new_designation)
                        # match_3 = re.search(r'^[A-Z]+\d*\s\d+\s[A-Z]+', new_designation)
                        # if match_3 and not verified:
                        #     print("MATCH 3")
                        #     #print(new_designation)
                        #     new_designation=match_3.group()
                        #     verified=True
                        # match_4 = re.match(r'^[A-Z][a-zA-Z\s]+?(?=\d)', new_designation)
                        # if match_4 and not verified:
                        #     print(" MATCH 4")
                        #     #print(new_designation)
                        #     new_designation=match_4.group()
                        #     verified=True

                        # match_5 = re.match(r'^[A-Za-z0-9\s\-]+', new_designation)
                        # if match_5:
                        #     #return match.group().strip()
                        #     new_designation=match_5.group().strip()
                        #     verified=True
                        #     #current_invoice['BLNo'].split()[2]]
                        
                        
                        print(new_designation)
                        if "BC OSV" in new_designation:
                            new_designation=new_designation.replace("BC OSV","OSV")
                        if "FR" in new_designation:
                            new_designation=new_designation.replace("FR","")

                        #print(no)
                        #print(f" {current_invoice['BLNo'].split()[2]} {new_designation}"


                    # Diviser la désignation en mots
                        mots = new_designation.strip().split()
                        last_single_match = None
                        for i in range(1, len(mots) + 1):
                            # Construire la chaîne de recherche avec les premiers i mots
                            search_string = ' '.join(mots[:i])
                            # Effectuer la recherche
                            matching_rows_temp = dataframe[dataframe['Nom'].str.contains(search_string, na=False, regex=False)]
            
                            # Vérification du nombre de correspondances
                            if len(matching_rows_temp) == 1:
                                # Mettre à jour le dernier résultat avec une seule correspondance
                                last_single_match = matching_rows_temp
                            elif len(matching_rows_temp) > 1:
                                # Mettre à jour le dernier résultat valide (plus d'une correspondance)
                                last_single_match = matching_rows_temp
                            # Continuer la boucle pour parcourir tous les mots

                        # Après la boucle, vérifier si un résultat unique a été trouvé
                        if last_single_match is not None:
                            matching_row = last_single_match.iloc[0]
                            Prix_Net = matching_row['Prix_Net']
                            Nom_produit = matching_row['Nom']
                        else:
                            # Si aucune correspondance unique n'a été trouvée, utiliser le dernier résultat valide
                            matching_row = last_single_match.iloc[0] if last_single_match is not None else None
                            if matching_row is not None:
                                Prix_Net = matching_row['Prix_Net']
                                Nom_produit = matching_row['Nom']
                            else:
                                # Aucune correspondance trouvée
                                Prix_Net = None
                                Nom_produit = None



                        
                        diff_ophtal=0

                        # try:
                        #         #dicimal_catalogue = Decimal(best_matching_row['Prix_Net']).quantize(Decimal('0.01'))
                        #     if discount=="0,00":
                        #         discount=discount.replace(",",".")
                        #         discount = Decimal(discount).quantize(Decimal('0.01'))
                        #         Prix_Net=str(Prix_Net).replace(",",".")
                        #         best_matching_prix_net = np.float64(Prix_Net)  # numpy.float64
                        #         best_matching_prix_net_decimal = Decimal(str(best_matching_prix_net))
                        #         diff_ophtal=best_matching_prix_net_decimal-discount
                        # except TypeError:
                        #         diff_ophtal = 0
                            

                                # Création du dictionnaire de produit
                            #new_designation = re.match(r'^[A-Z]+\s?\d+\s?[A-Z]+\s?\d+', description)
                        
                        try:
                            diff_ophtal = float(diff_ophtal)
                        except (ValueError, TypeError):
                            diff_ophtal = 0.0  # Set a default value if conversion fails
                        product = {
                                'Num': current_invoice['BLNo'].split()[2],        
                                'Description': description,
                                "New_Description" : new_designation,
                                'Discount': discount,
                                'Porteur': porteur,
                                'Date': current_invoice['BLNo'].split()[4],
                                'Qtt': qtt,
                                "Prix_Net": Prix_Net,
                                "Simulaire": Nom_produit,
                                "Diff": diff_ophtal
                        }
                        #print(product)
                            # Ajout du produit à la liste des produits de la facture courante
                    current_invoice['Products'].append(product)
        else:
            # Ajout de la ligne au champ Description du dernier produit de la facture courante
            if current_invoice.get('Products'):
                current_invoice['Products'][-1]['Description'] += ' ' + line.strip()
    print(no_match)
    # Ajout de la dernière facture à la liste des factures
    if current_invoice:
        invoices.append(current_invoice)
    
    return invoices



def decortiquer_ophtal_product(pdf_path):
    raw_text = extract_text_from_pdf(pdf_path)
    if not raw_text:
        return []
    structured_data = parse_invoice_text(raw_text)
    #print(structured_data)
    #print(type(structured_data))
    return structured_data




import os
import pandas as pd
from django.conf import settings



def database():
    #excel_file = "C:/Users/Amnzo/Desktop/ophtalmic/BL/DB.xlsx"
    data_dir = os.path.join(settings.BASE_DIR, 'comparateur', 'data')
    excel_file = os.path.join(data_dir, 'cata.xlsx')
    dfxl = pd.read_excel(excel_file, header=None)
    lignes_valides = []
    
    for index, ligne in dfxl.iterrows():
        #print(ligne)
        if not ligne.isnull().all():       
            nom = ligne[0]+ ' ' + ligne[1]
            valeur = ligne[2]            
            lignes_valides.append([nom, valeur])
    
    df_resultat = pd.DataFrame(lignes_valides, columns=['Nom','Prix_Net'])
    df_resultat['Prix_Net'] = df_resultat['Prix_Net'].astype(float)
    excel_file = os.path.join(data_dir, 'OUTPUT__Boite___.xlsx')
    df_resultat.to_excel(excel_file, index=False)



    #dfxl_output_boite = pd.read_excel(excel_file_output_boite, header=None)
    
    # Si vous avez besoin de traiter les données du deuxième fichier, faites-le ici
    # Par exemple, vous pouvez créer un DataFrame avec les données du deuxième fichier
    #df_output_boite = pd.DataFrame(dfxl_output_boite)
    
    # Vous pouvez également retourner les DataFrames si vous souhaitez les utiliser en dehors de cette fonction


    

from fuzzywuzzy import process,fuzz
def ophtal(request):
    if request.method == 'POST' and request.FILES['pdf_file']:
        #database()
        #print(data_tv1)
        #data_dir = os.path.join(settings.BASE_DIR, 'comparateur', 'data')
        #output_file = "output__.xlsx"
        #excel_file = os.path.join(data_dir, 'OUTPUT_____.xlsx')
        #dfxl = pd.read_excel(excel_file, header=None)
        
        #data_tv1.to_excel(excel_file, index=False)
        #print(data_tv1)
        pdf_file = request.FILES['pdf_file']
        temp_path = os.path.join(settings.MEDIA_ROOT, pdf_file.name)
        temp_path = os.path.join(settings.MEDIA_ROOT, pdf_file.name)
        with open(temp_path, 'wb+') as destination:
            for chunk in pdf_file.chunks():
                destination.write(chunk)
        structured_data = decortiquer_ophtal_product(temp_path)
        os.remove(temp_path)
        return render(request, 'excel/ophtal.html', {'formatted_commands': structured_data})
    return render(request, 'excel/ophtal.html')































# def decortiquer_ophtal(texte):
#     N_bl = None
#     lines = texte.split('\n')
#     for i, l in enumerate(lines):
#         if "B.L. No" in l:   
#             N_bl = l.split()[0]       
#     return N_bl


# def ophtal_old(request):
#     if request.method == 'POST' and request.FILES['pdf_file']:
#         excel_file = "C:/Users/Amnzo/Desktop/ophtalmic/BL/AMNZO.xlsx"
#         dataframe = pd.read_excel(excel_file)
#         # Example usage
#         pdf_path = 'C:/Users/Amnzo/Desktop/ophtalmic/FC22009537.pdf'
#         structured_data = decortiquer_ophtal_product(pdf_path)
#         #print(json.dumps(structured_data, indent=4))
  
#         dfs = []
#         path = "C:/Users/Amnzo/Desktop/ophtalmic/BL/"
#         filenames = ['BL_2021.csv', 'BL_2022.csv', 'BL_2023.csv']
#         for filename in filenames:
#             df = pd.read_csv(os.path.join(path, filename), encoding='latin1', delimiter=';')
#             dfs.append(df)

#         df_combined = pd.concat(dfs, ignore_index=True)
#         pdf_file = request.FILES['pdf_file']
#         pdf_data = pdf_file.read()
#         pdf_buffer = BytesIO(pdf_data)
#         pdf_reader = PdfReader(pdf_buffer)
#         content = ""
#         for page_num in range(len(pdf_reader.pages) - 1):
#             page = pdf_reader.pages[page_num]
#             text = page.extract_text()
#             encoded_text = text.encode('utf-8', 'ignore')  # Ignore characters that cannot be encoded
#             decoded_text = encoded_text.decode('utf-8')
#             content += decoded_text
        
#         commands = extract_commands_ophtal(content)
#         liste_cmds = []
#         for command in commands:
#             porteur = None
#             numero_bl = decortiquer_ophtal(command)
#             resultat = df_combined[df_combined['No livraison'] == numero_bl] 
#             if not resultat.empty: 
#                 for index, row in resultat.iterrows():       
#                     porteur = row.iloc[8]
#                     Date = row.iloc[10]
#                     designation = row.iloc[11]
#                     param = row.iloc[9] if not pd.isna(row.iloc[9]) else ''
#                     qtt = int(row.iloc[12])
#                     montant = row.iloc[3]
#                     best_match_name = None
#                     montant2 = Decimal(montant.replace(',', '.'))
#                     product_much=''
#                     # Find the closest matching product in the DataFrame
#                     best_match = process.extractOne(designation, dataframe['Nom'], scorer=fuzz.token_sort_ratio)
#                     #print(best_match)
#                     matching_row=None
#                     matching_row_nom="-"
#                     matching_row_price="-"
#                     if best_match[1] >= 80:  # Adjust threshold as needed
#                         product_much = best_match[0]
#                         matching_row = dataframe[dataframe['Nom'] == product_much].iloc[0]
#                         #print(f"-----{designation}-------")
#                         #print(matching_row['Nom'])
#                         matching_row_nom=matching_row['Nom']
#                         matching_row_price=matching_row['Prix_Net']
#                         #print(" ")
#                         #print(" ")
#                         #print(" ")

#                     formatted_ligne = {
#                         "numero_bl": f"{numero_bl}",
#                         "Porteur": porteur,
#                         "Date": Date,
#                         "Designation": f"{designation} {param}",
#                         'Qtt': qtt,
#                         "Montant_Facture": ((montant2)/2)/qtt,
#                         "search_term": designation,
#                         "Similaire": matching_row_nom,
#                         "Similaire_price": matching_row_price
#                     }
                    
#                     liste_cmds.append(formatted_ligne)
     
#         return render(request, 'excel/ophtal.html', {'formatted_commands': structured_data})
#     return render(request, 'excel/ophtal.html')



                    # try:
                    #     #dicimal_catalogue = Decimal(best_matching_prix_net).quantize(Decimal('0.01'))
                    #     if discount=="0,00":
                    #         discount=discount.replace(",",".")
                    #     discount = Decimal(discount).quantize(Decimal('0.01'))
                    #     #best_matching_prix_net = np.float64(10.50)  # numpy.float64
                    #     best_matching_prix_net_decimal = Decimal(str(best_matching_prix_net))
                    # except TypeError:
                    #     diff_ophtal = 0